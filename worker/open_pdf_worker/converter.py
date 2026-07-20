from __future__ import annotations

import re
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable

import camelot
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from open_pdf_worker.ocr import build_searchable_page_pdf
from open_pdf_worker.page_detection import classify_pages, pages_to_spec

ProgressCallback = Callable[[str, int, str], None]

CONFIDENCE_THRESHOLD = 0.85
ACCURACY_THRESHOLD = 95.0
MIN_CONFIDENCE_WHEN_ACCURATE = 0.75
REVIEW_SHEET_NAME = "Review"
OCR_REVIEW_THRESHOLD = 0.9418
STREAM_SUPPLEMENT_EDGE_TOL = 500

_MONEY_RE = re.compile(r"^[+-]?\$?\d{1,3}(?:,\d{3})*(?:\.\d{2})?$")
_DATE_RE = re.compile(r"\d{2}/\d{2}/\d{2}")
_POINTS_RE = re.compile(r"^\d{1,3}(?:,\d{3})+$")


@dataclass(frozen=True)
class WorksheetInfo:
    name: str
    source_pages: list[int]
    confidence: float
    extraction_method: str


@dataclass(frozen=True)
class ConversionResult:
    output_xlsx: Path
    worksheets: list[WorksheetInfo]
    warnings: list[str] = field(default_factory=list)


@dataclass(frozen=True)
class ExtractedTable:
    page: int
    order: int
    flavor: str
    confidence: float
    accuracy: float
    rows: list[list[str]]
    merges: list[tuple[int, int, int, int]]
    ocr_confidence: float | None = None


def _safe_sheet_name(base: str, used: set[str]) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in (" ", "_", "-") else "_" for ch in base)
    cleaned = cleaned.strip() or "Sheet"
    cleaned = cleaned[:31]
    candidate = cleaned
    suffix = 2
    while candidate in used:
        tail = f"_{suffix}"
        candidate = f"{cleaned[: 31 - len(tail)]}{tail}"
        suffix += 1
    used.add(candidate)
    return candidate


def _table_confidence(table: camelot.core.Table) -> float:
    report = table.parsing_report
    return float(report.get("confidence", 0))


def _table_accuracy(table: camelot.core.Table) -> float:
    report = table.parsing_report
    return float(report.get("accuracy", 0))


def _table_page(table: camelot.core.Table) -> int:
    return int(table.parsing_report.get("page", 1))


def _table_order(table: camelot.core.Table) -> int:
    return int(table.parsing_report.get("order", 1))


def _is_low_confidence(
    confidence: float,
    accuracy: float,
    *,
    ocr_confidence: float | None = None,
) -> bool:
    if ocr_confidence is not None and ocr_confidence < OCR_REVIEW_THRESHOLD:
        return True
    if accuracy >= ACCURACY_THRESHOLD and confidence >= MIN_CONFIDENCE_WHEN_ACCURATE:
        return False
    return confidence < CONFIDENCE_THRESHOLD or accuracy < ACCURACY_THRESHOLD


def _is_money_token(value: str) -> bool:
    text = value.strip().replace(" ", "")
    return bool(_MONEY_RE.fullmatch(text))


def _is_amount_token(value: str) -> bool:
    text = value.strip().replace(" ", "")
    return _is_money_token(text) or bool(_POINTS_RE.fullmatch(text))


def _split_stacked_value(value: str) -> tuple[str, str] | None:
    """Split Label\\n$amount (or reverse) into (label, amount)."""
    lines = [line.strip() for line in value.splitlines() if line.strip()]
    if len(lines) != 2:
        return None
    first, second = lines
    if _is_amount_token(first) and not _is_amount_token(second):
        return second, first
    if _is_amount_token(second) and not _is_amount_token(first):
        return first, second
    return None


def _expand_stacked_cells(rows: list[list[str]]) -> list[list[str]]:
    """Turn stacked label/amount cells into adjacent columns."""
    if not rows:
        return rows

    width = max(len(row) for row in rows)
    expanded: list[list[str]] = []
    for row in rows:
        padded = list(row) + [""] * (width - len(row))
        output: list[str] = []
        for index, cell in enumerate(padded):
            split = _split_stacked_value(cell)
            if split is None:
                output.append(cell)
                continue
            label, amount = split
            output.append(label)
            # Prefer an empty neighbor for the amount; otherwise insert a column.
            if index + 1 < len(padded) and not padded[index + 1].strip():
                padded[index + 1] = amount
            else:
                output.append(amount)
        expanded.append(output)

    max_width = max((len(row) for row in expanded), default=0)
    return [row + [""] * (max_width - len(row)) for row in expanded]


def _nonempty_cells(rows: list[list[str]]) -> list[str]:
    return [cell.strip() for row in rows for cell in row if cell and cell.strip()]


def _is_noise_table(rows: list[list[str]]) -> bool:
    """Drop prose/boilerplate tables that are not useful spreadsheet content."""
    cells = _nonempty_cells(rows)
    if len(cells) < 3:
        return False

    moneyish = sum(
        1
        for cell in cells
        if _is_money_token(cell) or any(_is_money_token(part) for part in cell.splitlines())
    )
    dateish = sum(1 for cell in cells if _DATE_RE.search(cell))
    sentence_like = sum(1 for cell in cells if len(cell.split()) >= 8)
    compact = sum(1 for cell in cells if len(cell.split()) <= 4)

    if moneyish == 0 and dateish == 0 and len(cells) <= 8:
        return True
    # Legal notices / policy pages: long sentences, no amounts worth exporting.
    if moneyish == 0 and sentence_like >= 3 and sentence_like >= 0.3 * len(cells):
        return True
    if sentence_like >= max(3, 0.45 * len(cells)) and moneyish <= 2 and dateish <= 1:
        return True
    if sentence_like >= 0.6 * len(cells) and compact < 0.25 * len(cells):
        return True
    return False


def _normalize_key(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip().lower()


def _row_content_keys(rows: list[list[str]]) -> set[str]:
    keys: set[str] = set()
    for row in rows:
        filled = [cell.strip() for cell in row if cell and cell.strip()]
        if not filled:
            continue
        keys.add(_normalize_key(" | ".join(filled)))
        for cell in filled:
            keys.add(_normalize_key(cell))
    return keys


def _label_amount_pairs(rows: list[list[str]]) -> dict[str, str]:
    pairs: dict[str, str] = {}
    for row in rows:
        cells = [cell.strip() for cell in row if cell and cell.strip()]
        if len(cells) < 2:
            continue
        for left, right in zip(cells, cells[1:]):
            if not _is_amount_token(left) and _is_amount_token(right):
                pairs[_normalize_key(left)] = right
            elif _is_amount_token(left) and not _is_amount_token(right):
                pairs[_normalize_key(right)] = left
    return pairs


def _fill_orphan_labels(rows: list[list[str]], pairs: dict[str, str]) -> list[list[str]]:
    """Fill trailing labels that lost their amount in a neighboring table on the same page."""
    if not pairs:
        return rows
    filled: list[list[str]] = []
    for row in rows:
        cells = list(row)
        nonempty = [(index, cell.strip()) for index, cell in enumerate(cells) if cell and cell.strip()]
        if len(nonempty) == 1:
            index, label = nonempty[0]
            amount = pairs.get(_normalize_key(label))
            if amount and not any(_is_amount_token(cell) for cell in cells):
                # Place amount in the next column when possible.
                if index + 1 < len(cells):
                    cells[index + 1] = amount
                else:
                    cells.append(amount)
        filled.append(cells)
    max_width = max((len(row) for row in filled), default=0)
    return [row + [""] * (max_width - len(row)) for row in filled]


def _supplemental_kv_rows(primary_keys: set[str], rows: list[list[str]]) -> list[list[str]]:
    """Keep compact label/amount rows from a looser pass that the primary pass missed."""
    extras: list[list[str]] = []
    for row in rows:
        cells = [cell.strip() for cell in row if cell and cell.strip()]
        if len(cells) < 2:
            continue
        if not any(_is_amount_token(cell) for cell in cells):
            continue
        # Transaction detail rows usually include dates; summary/fee rows do not.
        if any(_DATE_RE.search(cell) for cell in cells):
            continue
        if any(len(cell) > 80 for cell in cells):
            continue
        # Prefer short structured rows (fees, YTD, summary totals).
        if sum(len(cell.split()) for cell in cells) > 16:
            continue
        key = _normalize_key(" | ".join(cells))
        label_keys = {_normalize_key(cell) for cell in cells if not _is_amount_token(cell)}
        if key in primary_keys or (label_keys and label_keys <= primary_keys):
            continue
        extras.append(cells)
    return extras


def _read_tables(
    input_pdf: Path,
    page_spec: str,
    flavor: str,
    **camelot_kwargs: object,
) -> list[camelot.core.Table]:
    try:
        return list(
            camelot.read_pdf(
                str(input_pdf),
                pages=page_spec,
                flavor=flavor,
                **camelot_kwargs,
            )
        )
    except Exception as exc:
        # Surface extraction backend failures in diagnostics without aborting the other flavor.
        print(f"camelot {flavor} failed for {input_pdf} pages={page_spec!r}: {exc}", file=sys.stderr)
        return []


def _dataframe_to_rows(table: camelot.core.Table) -> list[list[str]]:
    dataframe = table.df.fillna("")
    rows = [[str(value).strip() for value in row] for row in dataframe.values.tolist()]
    return _expand_stacked_cells(rows)


def _row_fill_ratio(row: list[str]) -> float:
    if not row:
        return 0.0
    return sum(1 for value in row if value) / len(row)


def _trim_title_rows(rows: list[list[str]]) -> list[list[str]]:
    if len(rows) <= 1:
        return rows

    trimmed = list(rows)
    while len(trimmed) > 1:
        first = trimmed[0]
        second = trimmed[1]
        if _row_fill_ratio(first) <= 0.34 and _row_fill_ratio(second) >= 0.5:
            trimmed = trimmed[1:]
            continue
        break
    return trimmed


def _find_horizontal_merges(row_values: list[str], row_number: int) -> list[tuple[int, int, int, int]]:
    merges: list[tuple[int, int, int, int]] = []
    column = 0
    while column < len(row_values):
        value = row_values[column].strip()
        if not value:
            column += 1
            continue
        end_column = column
        while end_column + 1 < len(row_values) and not row_values[end_column + 1].strip():
            end_column += 1
        if end_column > column:
            merges.append((row_number, column + 1, row_number, end_column + 1))
        column = end_column + 1
    return merges


def _collect_merges(rows: list[list[str]]) -> list[tuple[int, int, int, int]]:
    merges: list[tuple[int, int, int, int]] = []
    for row_index, row in enumerate(rows, start=1):
        merges.extend(_find_horizontal_merges(row, row_index))
    return merges


def _select_tables(lattice_tables: list[camelot.core.Table], stream_tables: list[camelot.core.Table]) -> list[ExtractedTable]:
    all_tables = [*lattice_tables, *stream_tables]
    pages = sorted({_table_page(table) for table in all_tables})
    selected: list[ExtractedTable] = []

    for page in pages:
        page_lattice = [table for table in lattice_tables if _table_page(table) == page]
        page_stream = [table for table in stream_tables if _table_page(table) == page]

        page_tables: list[camelot.core.Table]
        flavor: str
        if page_lattice:
            page_tables = sorted(page_lattice, key=_table_order)
            flavor = "lattice"
        else:
            page_tables = sorted(page_stream, key=_table_order)
            flavor = "stream"

        page_rows: list[list[list[str]]] = []
        page_meta: list[tuple[int, float, float]] = []
        for table in page_tables:
            rows = _dataframe_to_rows(table)
            if flavor == "stream":
                rows = _trim_title_rows(rows)
            if not rows or not any(any(cell for cell in row) for row in rows):
                continue
            if _is_noise_table(rows):
                continue
            page_rows.append(rows)
            page_meta.append((_table_order(table), _table_confidence(table), _table_accuracy(table)))

        page_pairs: dict[str, str] = {}
        for rows in page_rows:
            page_pairs.update(_label_amount_pairs(rows))

        for rows, (order, confidence, accuracy) in zip(page_rows, page_meta):
            rows = _fill_orphan_labels(rows, page_pairs)
            selected.append(
                ExtractedTable(
                    page=page,
                    order=order,
                    flavor=flavor,
                    confidence=confidence,
                    accuracy=accuracy,
                    rows=rows,
                    merges=_collect_merges(rows),
                )
            )

    selected.sort(key=lambda table: (table.page, table.order))
    return selected


def _select_ocr_tables(
    lattice_tables: list[camelot.core.Table],
    stream_tables: list[camelot.core.Table],
    *,
    source_page: int,
    ocr_confidence: float,
) -> list[ExtractedTable]:
    selected = _select_tables(lattice_tables, stream_tables)
    adjusted: list[ExtractedTable] = []
    for table in selected:
        confidence = min(table.confidence, ocr_confidence)
        adjusted.append(
            ExtractedTable(
                page=source_page,
                order=table.order,
                flavor=f"ocr-{table.flavor}",
                confidence=confidence,
                accuracy=table.accuracy,
                rows=table.rows,
                merges=table.merges,
                ocr_confidence=ocr_confidence,
            )
        )
    return adjusted


def _merge_stream_supplement(
    primary: list[ExtractedTable],
    supplement_tables: list[camelot.core.Table],
) -> list[ExtractedTable]:
    """Add compact label/amount rows found only by a looser stream pass."""
    if not supplement_tables:
        return primary

    primary_keys: set[str] = set()
    max_order_by_page: dict[int, int] = {}
    for table in primary:
        primary_keys |= _row_content_keys(table.rows)
        max_order_by_page[table.page] = max(max_order_by_page.get(table.page, 0), table.order)

    extras: list[ExtractedTable] = []
    for page in sorted({_table_page(table) for table in supplement_tables}):
        page_supplement = [table for table in supplement_tables if _table_page(table) == page]
        page_extra_rows: list[list[str]] = []
        confidence = 0.0
        accuracy = 0.0
        for table in page_supplement:
            rows = _dataframe_to_rows(table)
            page_extra_rows.extend(_supplemental_kv_rows(primary_keys, rows))
            confidence = max(confidence, _table_confidence(table))
            accuracy = max(accuracy, _table_accuracy(table))

        # Deduplicate while preserving order.
        deduped: list[list[str]] = []
        seen: set[str] = set()
        for row in page_extra_rows:
            key = _normalize_key(" | ".join(row))
            if key in seen or key in primary_keys:
                continue
            seen.add(key)
            deduped.append(row)
            primary_keys.add(key)

        if not deduped:
            continue

        order = max_order_by_page.get(page, 0) + 1
        extras.append(
            ExtractedTable(
                page=page,
                order=order,
                flavor="stream-supplement",
                confidence=confidence,
                accuracy=accuracy,
                rows=deduped,
                merges=_collect_merges(deduped),
            )
        )

    merged = [*primary, *extras]
    merged.sort(key=lambda table: (table.page, table.order))
    return merged


def _extract_digital_tables(input_pdf: Path, page_spec: str) -> list[ExtractedTable]:
    lattice_tables = _read_tables(input_pdf, page_spec, "lattice")
    stream_tables = _read_tables(input_pdf, page_spec, "stream")
    selected = _select_tables(lattice_tables, stream_tables)
    supplement = _read_tables(
        input_pdf,
        page_spec,
        "stream",
        edge_tol=STREAM_SUPPLEMENT_EDGE_TOL,
    )
    return _merge_stream_supplement(selected, supplement)


def _extract_scan_tables(input_pdf: Path, scan_pages: list[int]) -> list[ExtractedTable]:
    extracted: list[ExtractedTable] = []
    for page_number in scan_pages:
        with tempfile.TemporaryDirectory(prefix="open-pdf-ocr-") as temp_dir:
            searchable_pdf = Path(temp_dir) / f"page-{page_number}.pdf"
            ocr_result = build_searchable_page_pdf(input_pdf, page_number, searchable_pdf)
            lattice_tables = _read_tables(searchable_pdf, "1", "lattice")
            stream_tables = _read_tables(searchable_pdf, "1", "stream")
            extracted.extend(
                _select_ocr_tables(
                    lattice_tables,
                    stream_tables,
                    source_page=page_number,
                    ocr_confidence=ocr_result.average_confidence,
                )
            )
    extracted.sort(key=lambda table: (table.page, table.order))
    return extracted


def _extract_tables(input_pdf: Path, page_spec: str) -> list[ExtractedTable]:
    digital_pages, scan_pages = classify_pages(input_pdf, page_spec)
    extracted: list[ExtractedTable] = []

    if digital_pages:
        extracted.extend(_extract_digital_tables(input_pdf, pages_to_spec(digital_pages)))
    if scan_pages:
        extracted.extend(_extract_scan_tables(input_pdf, scan_pages))

    extracted.sort(key=lambda table: (table.page, table.order))
    return extracted


def _write_text_cell(sheet: Worksheet, row: int, column: int, value: object) -> None:
    text = "" if value is None else str(value).strip()
    cell = sheet.cell(row=row, column=column, value=text)
    cell.number_format = "@"


def _write_table(sheet: Worksheet, rows: list[list[str]], merges: list[tuple[int, int, int, int]]) -> None:
    for row_index, row_values in enumerate(rows, start=1):
        for column_index, value in enumerate(row_values, start=1):
            _write_text_cell(sheet, row_index, column_index, value)

    for start_row, start_column, end_row, end_column in merges:
        if end_column > start_column or end_row > start_row:
            sheet.merge_cells(
                start_row=start_row,
                start_column=start_column,
                end_row=end_row,
                end_column=end_column,
            )


def _autosize_columns(sheet: Worksheet) -> None:
    for column_cells in sheet.columns:
        column_index = column_cells[0].column
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=0)
        sheet.column_dimensions[get_column_letter(column_index)].width = min(max(max_length + 2, 8), 40)


def _write_review_sheet(workbook: Workbook, entries: list[tuple[str, int, float, float, str]]) -> None:
    sheet = workbook.create_sheet(title=REVIEW_SHEET_NAME)
    headers = ["Worksheet", "Source page", "Confidence", "Accuracy", "Note"]
    for column_index, header in enumerate(headers, start=1):
        _write_text_cell(sheet, 1, column_index, header)

    for row_index, (name, page, confidence, accuracy, note) in enumerate(entries, start=2):
        _write_text_cell(sheet, row_index, 1, name)
        _write_text_cell(sheet, row_index, 2, str(page))
        _write_text_cell(sheet, row_index, 3, f"{confidence:.2f}")
        _write_text_cell(sheet, row_index, 4, f"{accuracy:.1f}")
        _write_text_cell(sheet, row_index, 5, note)

    _autosize_columns(sheet)


def convert_pdf_to_excel(
    input_pdf: Path,
    output_xlsx: Path,
    pages: str | None,
    on_progress: ProgressCallback,
) -> ConversionResult:
    on_progress("starting", 5, "Preparing conversion.")
    page_spec = pages or "all"

    on_progress("extracting", 20, f"Extracting tables from pages {page_spec}.")
    extracted_tables = _extract_tables(input_pdf, page_spec)
    if not extracted_tables:
        raise RuntimeError("No tables detected in the requested page range.")

    warnings: list[str] = []
    review_entries: list[tuple[str, int, float, float, str]] = []

    on_progress("writing", 70, f"Writing {len(extracted_tables)} worksheet(s).")
    workbook = Workbook()
    default_sheet = workbook.active
    assert default_sheet is not None
    workbook.remove(default_sheet)

    used_names: set[str] = set()
    worksheets: list[WorksheetInfo] = []

    for index, table in enumerate(extracted_tables, start=1):
        sheet_name = _safe_sheet_name(f"Table {index} p{table.page}", used_names)
        sheet = workbook.create_sheet(title=sheet_name)
        _write_table(sheet, table.rows, table.merges)

        metadata_row = len(table.rows) + 3
        _write_text_cell(sheet, metadata_row, 1, "Source page(s)")
        _write_text_cell(sheet, metadata_row, 2, str(table.page))
        _autosize_columns(sheet)

        worksheets.append(
            WorksheetInfo(
                name=sheet_name,
                source_pages=[table.page],
                confidence=table.confidence,
                extraction_method=table.flavor,
            )
        )

        if _is_low_confidence(
            table.confidence,
            table.accuracy,
            ocr_confidence=table.ocr_confidence,
        ):
            note = "Review extracted values against the source PDF."
            if table.flavor.startswith("ocr"):
                note = "Review OCR extraction against the source scan."
            warnings.append(f"{sheet_name} on page {table.page} needs review.")
            review_entries.append(
                (sheet_name, table.page, table.confidence, table.accuracy, note)
            )

    if review_entries:
        _write_review_sheet(workbook, review_entries)

    on_progress("finalizing", 90, "Saving workbook.")
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_xlsx)

    on_progress("complete", 100, "Conversion finished.")
    return ConversionResult(output_xlsx=output_xlsx, worksheets=worksheets, warnings=warnings)
