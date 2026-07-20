"""Shared subprocess harness and benchmark helpers for tests and scripts."""

from __future__ import annotations

import json
import resource
import subprocess
import sys
import threading
import time
from pathlib import Path

from openpyxl import load_workbook

from open_pdf_worker.converter import REVIEW_SHEET_NAME
from open_pdf_worker.version import PROTOCOL_VERSION

REPO_ROOT = Path(__file__).resolve().parents[2]
WORKER_ROOT = Path(__file__).resolve().parents[1]
CORPUS_DIR = REPO_ROOT / "corpus"
SAMPLE_PDF = CORPUS_DIR / "ruled_table.pdf"
EXPECTED_PATH = CORPUS_DIR / "ruled_table.expected.json"
MANIFEST_PATH = CORPUS_DIR / "manifest.json"
CORPUS_GENERATOR = CORPUS_DIR / "generate_digital_corpus.py"
SCAN_CORPUS_GENERATOR = CORPUS_DIR / "generate_scan_corpus.py"


def ensure_sample_pdf() -> Path:
    ensure_corpus()
    return SAMPLE_PDF


def ensure_corpus() -> None:
    ensure_digital_corpus()
    ensure_scan_corpus()


def ensure_digital_corpus() -> None:
    manifest = load_digital_corpus_manifest()
    missing_pdfs = [
        entry
        for entry in manifest
        if not (CORPUS_DIR / entry["pdf"]).exists()
    ]
    if missing_pdfs or not CORPUS_GENERATOR.exists():
        if not CORPUS_GENERATOR.exists():
            raise FileNotFoundError(f"Missing corpus generator: {CORPUS_GENERATOR}")
        subprocess.run(
            [sys.executable, str(CORPUS_GENERATOR)],
            check=True,
            cwd=REPO_ROOT,
        )


def ensure_scan_corpus() -> None:
    manifest = load_scan_corpus_manifest()
    missing_pdfs = [
        entry
        for entry in manifest
        if not (CORPUS_DIR / entry["pdf"]).exists()
    ]
    missing_expected = [
        entry
        for entry in manifest
        if not (CORPUS_DIR / entry["expected"]).exists()
    ]
    if missing_pdfs or missing_expected:
        if not SCAN_CORPUS_GENERATOR.exists():
            raise FileNotFoundError(f"Missing scan corpus generator: {SCAN_CORPUS_GENERATOR}")
        subprocess.run(
            [sys.executable, str(SCAN_CORPUS_GENERATOR)],
            check=True,
            cwd=REPO_ROOT,
        )


def load_digital_corpus_manifest() -> list[dict]:
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    return manifest["digital_classes"]


def load_scan_corpus_manifest() -> list[dict]:
    manifest = json.loads(MANIFEST_PATH.read_text(encoding="utf-8"))
    return manifest["scan_classes"]


def load_expected_spec(path: Path) -> dict:
    return json.loads(path.read_text(encoding="utf-8"))


def peak_memory_bytes() -> int:
    usage = resource.getrusage(resource.RUSAGE_CHILDREN)
    if sys.platform == "darwin":
        return usage.ru_maxrss
    return usage.ru_maxrss * 1024


def read_workbook_cells(workbook_path: Path) -> list[list[str]]:
    workbook = load_workbook(workbook_path, data_only=False)
    cells: list[list[str]] = []
    for sheet in workbook.worksheets:
        if sheet.title == REVIEW_SHEET_NAME:
            continue
        for row in sheet.iter_rows(values_only=True):
            cells.append(["" if value is None else str(value) for value in row])
    return cells


def read_workbook_cell_formats(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, data_only=False)
    formats: list[str] = []
    for sheet in workbook.worksheets:
        if sheet.title == REVIEW_SHEET_NAME:
            continue
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                if cell.value is not None and str(cell.value):
                    formats.append(cell.number_format)
    return formats


def workbook_sheet_names(workbook_path: Path) -> list[str]:
    workbook = load_workbook(workbook_path, data_only=False)
    return workbook.sheetnames


def workbook_has_review_sheet(workbook_path: Path) -> bool:
    return REVIEW_SHEET_NAME in workbook_sheet_names(workbook_path)


def workbook_merged_ranges(workbook_path: Path) -> dict[str, list[str]]:
    workbook = load_workbook(workbook_path, data_only=False)
    merged: dict[str, list[str]] = {}
    for sheet in workbook.worksheets:
        if sheet.title == REVIEW_SHEET_NAME:
            continue
        merged[sheet.title] = [str(item) for item in sheet.merged_cells.ranges]
    return merged


def workbook_source_pages(workbook_path: Path) -> list[int]:
    workbook = load_workbook(workbook_path, data_only=False)
    pages: list[int] = []
    for sheet in workbook.worksheets:
        if sheet.title == REVIEW_SHEET_NAME:
            continue
        for row in sheet.iter_rows(values_only=True):
            if row and row[0] == "Source page(s)" and row[1] is not None:
                pages.append(int(str(row[1])))
    return pages


def workbook_fingerprint(workbook_path: Path) -> list[tuple[str, list[tuple[int, int, str]]]]:
    workbook = load_workbook(workbook_path, data_only=False)
    fingerprint: list[tuple[str, list[tuple[int, int, str]]]] = []
    for sheet in workbook.worksheets:
        cells: list[tuple[int, int, str]] = []
        for row in sheet.iter_rows(min_row=1):
            for cell in row:
                if cell.value is not None and str(cell.value):
                    cells.append((cell.row, cell.column, str(cell.value)))
        fingerprint.append((sheet.title, cells))
    return fingerprint


def _normalize_cell(value: object) -> str:
    return str(value or "").replace("\n", " ").strip()


def flatten_values(cells: list[list[str]]) -> set[str]:
    normalized: set[str] = set()
    for row in cells:
        for value in row:
            text = _normalize_cell(value)
            if text:
                normalized.add(text)
    return normalized


def compare_expected_workbook(workbook_path: Path, expected: dict) -> dict:
    cells = read_workbook_cells(workbook_path)
    flat = flatten_values(cells)
    required = set(expected["required_values"])
    headers = set(expected.get("header_values", []))
    width = len(expected["rows"][0]) if expected.get("rows") else 0

    found_required = required & flat
    found_headers = headers & flat

    expected_rows = {tuple(_normalize_cell(value) for value in row) for row in expected.get("rows", [])}
    actual_rows: set[tuple[str, ...]] = set()
    if width:
        actual_rows = {
            tuple(_normalize_cell(value) for value in row[:width])
            for row in cells
            if len(row) >= width and any(row[:width])
        }
    matched_rows = len(expected_rows & actual_rows)

    recall = len(found_required) / len(required) if required else 1.0
    precision = len(found_required) / (len(flat) or 1)

    source_metadata_found = any(
        row and row[0] == "Source page(s)" for row in cells
    )
    review_sheet_found = workbook_has_review_sheet(workbook_path)
    expect_review_sheet = expected.get("expect_review_sheet", False)
    review_sheet_ok = review_sheet_found == expect_review_sheet

    merged_ranges = workbook_merged_ranges(workbook_path)
    merged_ok = True
    for spec in expected.get("merged_ranges", []):
        found = any(spec["range"] in ranges for ranges in merged_ranges.values())
        if not found:
            merged_ok = False
            break
        if spec.get("value"):
            value_found = spec["value"] in flat
            merged_ok = merged_ok and value_found

    source_pages = workbook_source_pages(workbook_path)
    expected_source_pages = expected.get("source_pages")
    source_pages_ok = (
        True
        if expected_source_pages is None
        else sorted(source_pages) == sorted(expected_source_pages)
    )

    return {
        "cell_recall": recall,
        "cell_precision": precision,
        "matched_rows": matched_rows,
        "expected_rows": len(expected_rows),
        "headers_found": len(found_headers),
        "headers_expected": len(headers),
        "source_metadata_found": source_metadata_found,
        "review_sheet_found": review_sheet_found,
        "review_sheet_ok": review_sheet_ok,
        "merged_ranges_ok": merged_ok,
        "source_pages_ok": source_pages_ok,
        "values_found": sorted(found_required),
        "values_missing": sorted(required - found_required),
    }


class WorkerClient:
    def __init__(self) -> None:
        self.process = subprocess.Popen(
            [sys.executable, "-m", "open_pdf_worker"],
            cwd=WORKER_ROOT,
            stdin=subprocess.PIPE,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
        )
        assert self.process.stdin is not None
        assert self.process.stdout is not None
        assert self.process.stderr is not None
        self._stderr_lines: list[str] = []
        self._stderr_thread = threading.Thread(target=self._drain_stderr, daemon=True)
        self._stderr_thread.start()

    def _drain_stderr(self) -> None:
        assert self.process.stderr is not None
        for line in self.process.stderr:
            self._stderr_lines.append(line)

    def send(self, message: dict) -> None:
        assert self.process.stdin is not None
        self.process.stdin.write(json.dumps(message) + "\n")
        self.process.stdin.flush()

    def read_event(self) -> dict:
        assert self.process.stdout is not None
        line = self.process.stdout.readline()
        if not line:
            raise RuntimeError("Worker closed stdout unexpectedly.")
        return json.loads(line)

    def close(self) -> list[str]:
        assert self.process.stdin is not None
        self.process.stdin.close()
        self.process.wait(timeout=30)
        self._stderr_thread.join(timeout=5)
        return list(self._stderr_lines)


def handshake(client: WorkerClient) -> dict:
    client.send({"type": "handshake", "protocol_version": PROTOCOL_VERSION})
    event = client.read_event()
    if event["type"] != "handshake_ack":
        raise RuntimeError(event)
    return event


def run_conversion(
    input_pdf: Path,
    output_xlsx: Path,
    *,
    pages: str | None = "1",
    request_id: str = "convert",
) -> tuple[list[dict], float, int, int]:
    if output_xlsx.exists():
        output_xlsx.unlink()

    client = WorkerClient()
    messages = [
        {"type": "handshake", "protocol_version": PROTOCOL_VERSION},
        {
            "type": "convert",
            "request_id": request_id,
            "input_pdf": str(input_pdf.resolve()),
            "output_xlsx": str(output_xlsx.resolve()),
            **({"pages": pages} if pages is not None else {}),
        },
    ]

    started = time.perf_counter()
    for message in messages:
        client.send(message)

    events: list[dict] = []
    while True:
        event = client.read_event()
        events.append(event)
        if event["type"] in {"complete", "error"}:
            break

    client.close()
    elapsed = time.perf_counter() - started
    return events, elapsed, peak_memory_bytes(), client.process.returncode or 0


def benchmark_digital_corpus(tmp_dir: Path) -> dict:
    ensure_corpus()
    manifest = load_digital_corpus_manifest()
    by_class: dict[str, dict] = {}

    for entry in manifest:
        pdf_path = CORPUS_DIR / entry["pdf"]
        expected_path = CORPUS_DIR / entry["expected"]
        expected = load_expected_spec(expected_path)
        output = tmp_dir / f"{entry['id']}.xlsx"
        events, runtime_seconds, peak_memory, exit_code = run_conversion(
            pdf_path,
            output,
            pages=entry.get("pages", "1"),
            request_id=f"benchmark-{entry['id']}",
        )
        metrics = compare_expected_workbook(output, expected)
        metrics.update(
            {
                "document_class": entry["id"],
                "document": str(pdf_path),
                "runtime_seconds": round(runtime_seconds, 3),
                "peak_memory_bytes": peak_memory,
                "exit_code": exit_code,
                "progress_events": sum(1 for event in events if event["type"] == "progress"),
                "completed": events[-1]["type"] == "complete" if events else False,
            }
        )
        by_class[entry["id"]] = metrics

    return {"by_class": by_class}


def benchmark_scan_corpus(tmp_dir: Path) -> dict:
    ensure_scan_corpus()
    manifest = load_scan_corpus_manifest()
    by_class: dict[str, dict] = {}

    for entry in manifest:
        pdf_path = CORPUS_DIR / entry["pdf"]
        expected_path = CORPUS_DIR / entry["expected"]
        expected = load_expected_spec(expected_path)
        output = tmp_dir / f"{entry['id']}.xlsx"
        events, runtime_seconds, peak_memory, exit_code = run_conversion(
            pdf_path,
            output,
            pages=entry.get("pages", "1"),
            request_id=f"scan-benchmark-{entry['id']}",
        )
        metrics = compare_expected_workbook(output, expected)
        extraction_methods = []
        if events and events[-1]["type"] == "complete":
            extraction_methods = [
                worksheet.get("extraction_method", "")
                for worksheet in events[-1].get("worksheets", [])
            ]
        metrics.update(
            {
                "document_class": entry["id"],
                "document": str(pdf_path),
                "runtime_seconds": round(runtime_seconds, 3),
                "peak_memory_bytes": peak_memory,
                "exit_code": exit_code,
                "progress_events": sum(1 for event in events if event["type"] == "progress"),
                "completed": events[-1]["type"] == "complete" if events else False,
                "extraction_methods": extraction_methods,
            }
        )
        by_class[entry["id"]] = metrics

    return {"by_class": by_class}
